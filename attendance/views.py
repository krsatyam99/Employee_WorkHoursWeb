from django.shortcuts import get_object_or_404, redirect, render
from attendance.models import employee_details,AttendanceTransaction
from django.http import JsonResponse,HttpResponse
from django.db.models import Count

from django.db.models import F
from datetime import date



import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill



# Create your views here.
def testing( request):
    if request.method =='GET':
        return render(request, 'base.html')
    else:
        pass




def employee_info(request,id =None):
    if request.method == 'GET':
        context={}
        info = employee_details.objects.all().order_by('Hire_Date')
        context['designation'] = employee_details.DESIGNATION_CHOICES
        context['department'] = employee_details.DEPARTMENT_CHOICES

       
        context['info']=info
        return render(request,'update.html',context,)


# views.py

from django.shortcuts import render
from .models import employee_details

def create(request):
    context = {}
    context['designation'] = employee_details.DESIGNATION_CHOICES
    context['department'] = employee_details.DEPARTMENT_CHOICES

    if request.method == "POST":
        # Handle form submission
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        hire_date = request.POST.get('hire_date')
        designation = request.POST.get('designation')
        department = request.POST.get('department')

        employee = employee_details(
            First_Name=first_name,
            Last_Name=last_name,
            Hire_Date=hire_date,
            Designation=designation,
            Department=department
        )
        employee.save()
        return redirect('employe_info')
    else:
        
        return render(request, 'add_employee.html', context)

def update(request, id):
    context = {}
    context['designation'] = employee_details.DESIGNATION_CHOICES
    context['department'] = employee_details.DEPARTMENT_CHOICES

    if request.method == "POST":
    
            
        

        employee = get_object_or_404(employee_details, id=id)
        employee.First_Name = request.POST.get('first_name')
        employee.Last_Name = request.POST.get('last_name')
        employee.Department =request.POST.get('department')
        
        employee.Hire_Date = request.POST.get('hire_date')
        employee.Designation = request.POST.get('designation')
        employee.save()
        return redirect('employe_info')
    else:
        
        return redirect('employe_info')

       




def delete_employee(request,id):

    info = employee_details.objects.filter(id=id)
    info.delete()
    return redirect('employe_info')






def attendance_report(request):
    employee_ids = AttendanceTransaction.objects.values_list('employee_id', flat=True).distinct()
    employee_data = {}

    for employee_id in employee_ids:
        days_worked_data = (
            AttendanceTransaction.objects
            .filter(employee_id=employee_id)
            .values('employee__Full_Name', 'employee__Employee_Code', 'first_punch_date')
            .annotate(total_occurrences=Count('first_punch_date'))
        )

        total_occurrences = sum(entry['total_occurrences'] for entry in days_worked_data)

        print(f'Total days worked for employee {employee_id}: {total_occurrences} occurrences')

        employee_data[employee_id] = {
            'Full_Name': days_worked_data[0]['employee__Full_Name'],
            'Employee_Code': days_worked_data[0]['employee__Employee_Code'],
            'total_occurrences': total_occurrences,
            'days_worked_data': days_worked_data,
        }

    all_attendance_data = AttendanceTransaction.objects.all()

    context = {
        'employee_data': employee_data,
        'all_attendance_data': all_attendance_data,
    }

    return render(request, 'attendance_report.html', context)







def generate_and_download_excel(request):
    # Fetch data
    employee_ids = AttendanceTransaction.objects.values_list('employee_id', flat=True).distinct()
    employee_data = {}

    for employee_id in employee_ids:
        days_worked_data = (
            AttendanceTransaction.objects
            .filter(employee_id=employee_id)
            .values('employee__Full_Name', 'employee__Employee_Code', 'first_punch_date')
            .annotate(total_occurrences=Count('first_punch_date'))
        )

        total_occurrences = sum(entry['total_occurrences'] for entry in days_worked_data)

        print(f'Total days worked for employee {employee_id}: {total_occurrences} occurrences')

        employee_data[employee_id] = {
            'Full_Name': days_worked_data[0]['employee__Full_Name'],
            'Employee_Code': days_worked_data[0]['employee__Employee_Code'],
            'total_occurrences': total_occurrences,
            'days_worked_data': days_worked_data,
        }

    # Generate Excel
    wb = openpyxl.Workbook()
    ws = wb.active

    # Merge cells for the title
    ws.merge_cells('A1:D4')

    # Add title in the center
    title_cell = ws['A1']
    title_cell.value = 'Attendance Report'
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(size=14, bold=True, color='000000')

   # Set header row with black text color and yellow cell color
    header_row = ws['A5:D5']
    header_labels = ['Employee ID', 'Full Name', 'Employee Code', 'Total Attendance']

    for idx, cell in enumerate(header_row[0]):
        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        cell.font = Font(color='000000', bold=True)
        cell.value = header_labels[idx]

    # Adjust column widths for better spacing
    ws.column_dimensions['A'].width = 15  # Employee ID
    ws.column_dimensions['B'].width = 30  # Full Name
    ws.column_dimensions['C'].width = 15  # Employee Code
    ws.column_dimensions['D'].width = 20  # Total Attendance

    # Populate data
    for row_num, (employee_id, data) in enumerate(employee_data.items(), start=6):
        ws.cell(row=row_num, column=1, value=employee_id)
        ws.cell(row=row_num, column=2, value=data['Full_Name'])
        ws.cell(row=row_num, column=3, value=data['Employee_Code'])
        ws.cell(row=row_num, column=4, value=data['total_occurrences'])

    # Save the workbook
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=attendance_report.xlsx'
    wb.save(response)

    return response













def get_attendance_data(request):
    today = date.today()
    formatted_today = today.strftime('%d/%m/%Y')

    # Query to get the required information for the present date
    attendance_data = AttendanceTransaction.objects.filter(first_punch_date=today).values(
        employee_code=F('employee__Employee_Code'),
        full_name=F('employee__Full_Name'),
        first_punch_time_annotation=F('first_punch_time'),  
        last_punch_time_value=F('last_punch_time'),
        total_hour_working_value=F('total_hour_working')
    )

    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active

    # Title cell
    title_cell = ws['A1']
    title_cell.value = f'Attendance Report {formatted_today}'
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.font = Font(size=14, bold=True, color='000000')
    ws.merge_cells('A1:F1')  # Merge cells for the title

    
    # Header row
    header_row = ws['A5:E5']
    header_labels = ['Employee Code', 'Full Name',  'In time', 'Out time','Duration']

    for idx, cell in enumerate(header_row[0]):
        cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')  # Yellow fill
        cell.font = Font(color='000000', bold=True, size=12)  # Black text, bold, font size 12
        cell.value = header_labels[idx]


    # Adjust column widths for better spacing
    ws.column_dimensions['A'].width = 15  # Employee Code
    ws.column_dimensions['B'].width = 30  # Full Name
    ws.column_dimensions['C'].width = 20  # First Punch
    ws.column_dimensions['D'].width = 20  # Last Punch
    ws.column_dimensions['E'].width = 25  # Total Hour Worked Today

 

    # Add data to the worksheet
    for entry in attendance_data:
        data_row = [
            entry['employee_code'],
            entry['full_name'],
            entry['first_punch_time_annotation'],
            entry['last_punch_time_value'],
            entry['total_hour_working_value']
        ]
        ws.append(data_row)

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=attendance_data.xlsx'
    wb.save(response)

    return response
